import streamlit as st
import requests
from bs4 import BeautifulSoup
import json
import pandas as pd
from datetime import datetime
import base64
from ics import Calendar, Event
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import timezone
import pytz

st.set_page_config(page_title="Sugarfitness-crawler 💪", page_icon=":emoji:")
st.title("Sugarfitness-crawler 💪")

def create_calendar_entry(session_detail):
    event = Event()
    event.name = session_detail['Class Title']
    # Get the current timezone
    current_tz = pytz.timezone('Europe/Budapest')
    # Parse the date and time strings and make them timezone-aware
    event.begin = current_tz.localize(datetime.strptime(session_detail['Scheduled Date'] + ' ' + session_detail['Start Time'], '%Y/%m/%d %A %H:%M'))
    event.end = current_tz.localize(datetime.strptime(session_detail['Scheduled Date'] + ' ' + session_detail['End Time'], '%Y/%m/%d %A %H:%M'))
    event.description = f"Trainer: {session_detail['Trainer Name']}"
    event.location = session_detail['Class Location']  # Add the location to the event
    return event

def get_calendar_download_link(session_details):
    calendar = Calendar(events=[create_calendar_entry(session_detail) for session_detail in session_details])
    b64 = base64.b64encode(calendar.serialize().encode())  # Use serialize() instead of str()
    return f'<a href="data:text/calendar;base64,{b64.decode()}" download="calendar.ics">Download Calendar file</a>'

def fetch_data(url):
    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        script_tag = soup.find('script', type='application/json')

        if script_tag:
            json_string = script_tag.string
            data = json.loads(json_string)
            sessions = data['props']['pageProps']['sessions']

            session_details = []
            for session in sessions:
                class_info = session.get('class', {})
                trainer_info = session.get('trainer', {})
                max_headcount = session.get('max_headcount')
                current_headcount = session.get('current_headcount')
                location = session.get('location', 'Unknown')  # Directly get the location
                if isinstance(location, dict):
                    location = location.get('title', 'Unknown')  # Extract the 'title' from the location dictionary
                available_slots = max_headcount - current_headcount
                available_slots = 'NO AVAILABLE SLOTS' if available_slots == 0 else str(available_slots)  # Convert to string
                start_time = datetime.fromisoformat(session.get('start'))
                end_time = datetime.fromisoformat(session.get('end'))
                length_of_class = f"{int((end_time - start_time).total_seconds() / 60)} mins"
                session_detail = {
                    'Class Title': class_info.get('title'),
                    'Class Difficulty': class_info.get('difficulty'),
                    'Class Category': class_info.get('category'),
                    'Class Location': location,  # Use the location directly
                    'Scheduled Date': datetime.fromisoformat(session.get('date')).strftime('%Y/%m/%d %A'),
                    'Start Time': start_time.strftime('%H:%M'),
                    'End Time': end_time.strftime('%H:%M'),
                    'Length of Class': length_of_class,
                    'Headcount': f"{max_headcount}/{current_headcount}@{datetime.now().strftime('%m-%d')}",
                    'Available Slots': available_slots,
                    'Trainer Name': f"{trainer_info.get('last_name', '')} {trainer_info.get('first_name', '')}",
                    'Trainer Gender': trainer_info.get('gender'),
                    'Trainer Position': trainer_info.get('position'),
                }
                session_details.append(session_detail)

            df = pd.DataFrame(session_details)
            df.columns = df.columns.astype(str)  # Ensure all column headers are strings

            return df
        else:
            st.error("Could not find the JSON data in the HTML source.")
    else:
        st.error('Failed to retrieve the page')

def to_excel(df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    tab = Table(displayName="Table1", ref=f"A1:{chr(65 + df.shape[1])}{df.shape[0] + 1}")

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    wb.save(output)
    return output.getvalue()

def get_table_download_link(df):
    val = to_excel(df)
    b64 = base64.b64encode(val)
    return f'<a href="data:application/octet-stream;base64,{b64.decode()}" download="extract.xlsx">Download Excel file</a>'

url = 'https://www.sugarfitness.hu/'
df = fetch_data(url)

if df is not None:
    unique_dates = df['Scheduled Date'].unique()
    unique_trainers = df['Trainer Name'].unique()
    unique_class_titles = df['Class Title'].unique()

    selected_date = st.selectbox('Select a date', options=[''] + list(unique_dates))
    selected_trainer = st.selectbox('Select a trainer', options=[''] + list(unique_trainers))
    selected_class_title = st.selectbox('Select a class title', options=[''] + list(unique_class_titles))

    if st.button('Clear all filters'):
        selected_date = ''
        selected_trainer = ''
        selected_class_title = ''

    filtered_df = df.copy()
    if selected_date:
        filtered_df = filtered_df[filtered_df['Scheduled Date'] == selected_date]
    if selected_trainer:
        filtered_df = filtered_df[filtered_df['Trainer Name'] == selected_trainer]
    if selected_class_title:
        filtered_df = filtered_df[filtered_df['Class Title'] == selected_class_title]

    st.dataframe(filtered_df)

    st.markdown(get_table_download_link(df), unsafe_allow_html=True)

    st.markdown(get_calendar_download_link(filtered_df.to_dict('records')), unsafe_allow_html=True)